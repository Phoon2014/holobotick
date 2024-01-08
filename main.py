import telebot
import openpyxl
from telebot import types
excel_file_path = 'hol.xlsx'
workbook = openpyxl.load_workbook(excel_file_path)
sheet = workbook.active
token = '6696612926:AAEYYUvxbvLm4ol9u6fSnaW1nVpyfyMxMBQ'
bot = telebot.TeleBot(token)
user_selections = {}
products = [
    "Классик (черный)", "Зеро (белый)", "Беспонтовый ГеймФулл (Синий)", "Ягоды (красный)",
    "Классик (зеленый)", "Зеро белый", "Голубика (синий)", "Манго (желтый)", "Кокос + ягоды (голубой)",
    "Оригинальный", "Манго", "Яблоко Киви", "Классик черный", "Голубика Гранат (фиолетовый)",
    "Манго Лайм (зеленый)", "Апельсин Маракуя", "Эктив", "Бабл Гам", "Кокос", "Малина", "Манго (синий)", "Экстра",
    "Ананас", "Апельсин", "Вода", "Зеленая классика", "Обожаю Персик", "Черный с лимоном", "Земляничка и клюковка",
    "Кола", "Кола Зеро", "Пепси", "7 UP", "FANTA", "Боржоми", "Сникерс", "Марс", "Твикс", "Баунти", "Маршмелоу",
    "Кофе Гуарана", "Орехи Манго", "Печенье"
]
unique_products = [f"{product} ({index})" for index, product in enumerate(products, start=1)]
print(unique_products)
@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    if message.text == '/start':
        bot.reply_to(message, "Привет! Я чат-бот 😇 Что ты съел сегодня?", reply_markup=main_menu_markup)
    elif message.text == '/help':
        bot.reply_to(message, "Помощь и инструкции...")
def hide_keyboard(message):
    markup = types.ReplyKeyboardRemove(selective=False)
    bot.send_message(message.chat.id, "Главное меню", reply_markup=markup)
# Словарь с соответствиями оригинальных названий продуктов и псевдонимов
product_aliases = {
    "Зеленая классика": "Зеленый классический",
    "Обожаю Персик": "Персик",
    "Черный с лимоном": "Лимон",
    "Земляничка и клюковка": "Зеленый с ягодами",
    # Добавьте другие продукты и их псевдонимы
}


def get_user_history(user_row):
    username = sheet.cell(row=user_row, column=1).value
    history = {"Энергетики": {}, "LIPTON": {}, "Лимонад": {}, "Снеки": {}, "Другие": {}}

    for col in range(2, sheet.max_column + 1):
        product_count = sheet.cell(row=user_row, column=col).value
        if product_count:
            product_name = products[col - 2]

            # Используем псевдоним, если он определен
            product_alias = product_aliases.get(product_name, product_name)

            category = categorize_product(product_name)

            # Проверяем, есть ли уже такая категория в истории, если нет - добавляем
            if category not in history:
                history[category] = {}

            history[category][product_alias] = product_count

    return history


def categorize_product(product_name):
    energy_categories = {
        "Адреналин": ["Классик (черный)", "Зеро (белый)", "Беспонтовый ГеймФулл (Синий)", "Ягоды (красный)"],
        "LIT Energy": ["Классик (зеленый)", "Зеро белый", "Голубика (синий)", "Манго (желтый)", "Кокос + ягоды (голубой)"],
        "CYBERWATER": ["Оригинальный", "Манго", "Яблоко Киви"],
        "VOLT": ["Классик черный", "Манго Лайм (зеленый)", "Апельсин Маракуя", "Голубика Гранат (фиолетовый)"],
        "TORNADO": ["Эктив", "Бабл Гам", "Кокос", "Малина"],
        "GORILLA": ["Манго (синий)", "Экстра", "Ананас", "Апельсин"]
    }

    lipton_categories = {
        "LIPTON": ["Зеленая классика", "Обожаю Персик", "Черный с лимоном", "Земляничка и клюковка"]
    }

    lemonade_categories = {
        "Лимонад": ["Кола", "Кола Зеро", "Пепси", "7 UP", "FANTA"]
    }

    snack_categories = {
        "Снеки": ["Сникерс", "Марс", "Твикс", "Баунти", "Маршмелоу", "Кофе Гуарана", "Орехи Манго", "Печенье"]
    }

    for category, products_in_category in energy_categories.items():
        if product_name in products_in_category:
            return category

    for category, products_in_category in lipton_categories.items():
        if product_name in products_in_category:
            return category

    for category, products_in_category in lemonade_categories.items():
        if product_name in products_in_category:
            return category

    for category, products_in_category in snack_categories.items():
        if product_name in products_in_category:
            return category

    return "Другие"


category_aliases = {
    "Зеленая классика": "Зеленый классический",
    "Обожаю Персик": "Персик",
    "Черный с лимоном": "Черный с лимоном",
    "Земляничка и клюковка": "Зеленый с ягодами",
}

@bot.message_handler(func=lambda message: message.text == "История", content_types=['text'])
def send_user_history(message):
    username = message.from_user.username
    user_row = find_user_row(username)

    if user_row is not None:
        user_history = get_user_history(user_row)
        formatted_history = format_user_history(user_history)
        bot.reply_to(message, f"Ваша история выбора продуктов:\n\n{formatted_history}")
    else:
        bot.reply_to(message, f"Пользователь не найден в таблице. Пожалуйста, выполните /start для начала.")

    send_welcome_with_menu(message)


def format_user_history(user_history):
    formatted_history = []
    for category, products_in_category in user_history.items():
        for product, count in products_in_category.items():
            formatted_history.append(f"{category}: {product} ({count})")

    return '\n'.join(formatted_history)
@bot.message_handler(func=lambda message: message.text == "История", content_types=['text'])
def send_user_history(message):
    username = message.from_user.username
    user_row = find_user_row(username)

    if user_row is not None:
        user_history = get_user_history(user_row)
        bot.reply_to(message, f"Ваша история выбора продуктов:\n\n{user_history}")
    else:
        bot.reply_to(message, f"Пользователь не найден в таблице. Пожалуйста, выполните /start для начала.")


    send_welcome(message)

main_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
main_menu_markup.add(types.KeyboardButton("Энергетики"),
                     types.KeyboardButton("Вода, чай и лимонад"),
                     types.KeyboardButton("Батончики и снеки"),
                     types.KeyboardButton("История"))

energy_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
energy_menu_markup.add(types.KeyboardButton("Адреналин"),
                       types.KeyboardButton("ЛИТ Энерджи"),
                       types.KeyboardButton("CYBERWATER"),
                       types.KeyboardButton("VOLT"),
                       types.KeyboardButton("TORNADO"),
                       types.KeyboardButton("GORILLA"))

adrenalin_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
adrenalin_menu_markup.add(types.KeyboardButton("Классик (черный)"),
                          types.KeyboardButton("Зеро (белый)"),
                          types.KeyboardButton("Беспонтовый ГеймФулл (Синий)"),
                          types.KeyboardButton("Ягоды (красный)"))

lit_energy_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
lit_energy_menu_markup.add(types.KeyboardButton("Классик (зеленый)"),
                          types.KeyboardButton("Зеро белый"),
                          types.KeyboardButton("Голубика (синий)"),
                          types.KeyboardButton("Манго (желтый)"),
                          types.KeyboardButton("Кокос + ягоды (голубой)"))

cyberwater_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
cyberwater_menu_markup.add(types.KeyboardButton("Оригинальный"),
                           types.KeyboardButton("Манго"),
                           types.KeyboardButton("Яблоко Киви"))

volt_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
volt_menu_markup.add(types.KeyboardButton("Классик черный"),
                    types.KeyboardButton("Манго Лайм (зеленый)"),
                    types.KeyboardButton("Апельсин Маракуя"),
                    types.KeyboardButton("Голубика Гранат (фиолетовый)"))

tornado_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
tornado_menu_markup.add(types.KeyboardButton("Эктив"),
                       types.KeyboardButton("Бабл Гам"),
                       types.KeyboardButton("Кокос"),
                       types.KeyboardButton("Малина"))

gorilla_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
gorilla_menu_markup.add(types.KeyboardButton("Манго (синий)"),
                       types.KeyboardButton("Экстра"),
                       types.KeyboardButton("Ананас"),
                       types.KeyboardButton("Апельсин"))

lipton_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
lipton_menu_markup.add(types.KeyboardButton("Зеленая классика"),
                      types.KeyboardButton("Обожаю Персик"),
                      types.KeyboardButton("Черный с лимоном"),
                      types.KeyboardButton("Земляничка и клюковка"))

lemonade_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
lemonade_menu_markup.add(types.KeyboardButton("Кола"),
                        types.KeyboardButton("Кола Зеро"),
                        types.KeyboardButton("Пепси"),
                        types.KeyboardButton("7 UP"),
                        types.KeyboardButton("FANTA"))

water_tea_lemonade_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
water_tea_lemonade_menu_markup.add(types.KeyboardButton("Вода"),
                                   types.KeyboardButton("Боржоми"),
                                   types.KeyboardButton("Чай LIPTON"),
                                   types.KeyboardButton("ЛИМОНАД"))

snacks_menu_markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
snacks_menu_markup.add(types.KeyboardButton("Сникерс"), types.KeyboardButton("Марс"),
                       types.KeyboardButton("Твикс"), types.KeyboardButton("Баунти"),
                       types.KeyboardButton("Маршмелоу"),
                       types.KeyboardButton("Кофе Гуарана",),
                       types.KeyboardButton("Орехи Манго"),
                       types.KeyboardButton("Печенье"))

@bot.message_handler(func=lambda message: message.text == "Батончики и снеки", content_types=['text'])
def send_snacks_menu(message):
    bot.reply_to(message, "Что схомячил, а?", reply_markup=snacks_menu_markup)

@bot.message_handler(func=lambda message: message.text == "Энергетики")
def send_energy_menu(message):
    bot.reply_to(message, "Что заставляет биться твое сердце чаще?", reply_markup=energy_menu_markup)

@bot.message_handler(func=lambda message: message.text == "Вода, чай и лимонад", content_types=['text'])
def send_water_tea_lemonade_menu(message):
    bot.reply_to(message, "Чем ты освежился, приятель?", reply_markup=water_tea_lemonade_menu_markup)

@bot.message_handler(func=lambda message: message.text in ["Вода", "Боржоми"], content_types=['text'])
def handle_water_borjomi_selection(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text == "Чай LIPTON", content_types=['text'])
def send_lipton_menu(message):
    bot.reply_to(message, "Какой чай ты любишь пить с бабулей?", reply_markup=lipton_menu_markup)

@bot.message_handler(func=lambda message: message.text == "ЛИМОНАД", content_types=['text'])
def send_lemonade_menu(message):
    bot.reply_to(message, "Что не дало тебе умереть от жажды?", reply_markup=lemonade_menu_markup)

@bot.message_handler(func=lambda message: message.text in [
    "Зеленая классика", "Обожаю Персик",
    "Черный с лимоном", "Земляничка и клюковка"
], content_types=['text'])
def handle_lipton_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Сникерс", "Марс", "Твикс", "Баунти", "Маршмелоу",
    "Кофе Гуарана", "Орехи Манго", "Печенье"
])
def handle_snack_selection(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Кола", "Кола Зеро", "Пепси", "7 UP", "FANTA"
], content_types=['text'])
def handle_lemonade_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text == "Энергетики", content_types=['text'])
def send_energy_menu(message):
    bot.reply_to(message, "Что заставляет биться твое сердце чаще?", reply_markup=energy_menu_markup)

@bot.message_handler(func=lambda message: message.text == "Адреналин", content_types=['text'])
def send_adrenalin_menu(message):
    bot.reply_to(message, "Выберите вкус Адреналина:", reply_markup=adrenalin_menu_markup)

@bot.message_handler(func=lambda message: message.text == "ЛИТ Энерджи", content_types=['text'])
def send_lit_energy_menu(message):
    bot.reply_to(message, "Выберите вкус LIT Energy:", reply_markup=lit_energy_menu_markup)

@bot.message_handler(func=lambda message: message.text == "CYBERWATER", content_types=['text'])
def send_cyberwater_menu(message):
    bot.reply_to(message, "Выберите вкус CYBERWATER:", reply_markup=cyberwater_menu_markup)

@bot.message_handler(func=lambda message: message.text == "VOLT", content_types=['text'])
def send_volt_menu(message):
    bot.reply_to(message, "Выберите вкус VOLT:", reply_markup=volt_menu_markup)

@bot.message_handler(func=lambda message: message.text == "TORNADO", content_types=['text'])
def send_tornado_menu(message):
    bot.reply_to(message, "Выберите вкус TORNADO:", reply_markup=tornado_menu_markup)

@bot.message_handler(func=lambda message: message.text == "GORILLA", content_types=['text'])
def send_gorilla_menu(message):
    bot.reply_to(message, "Выберите вкус GORILLA:", reply_markup=gorilla_menu_markup)

@bot.message_handler(func=lambda message: message.text in [
    "Классик (черный)", "Зеро (белый)",
    "Беспонтовый ГеймФулл (Синий)", "Ягоды (красный)"
], content_types=['text'])
def handle_adrenalin_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Классик (зеленый)", "Зеро белый",
    "Голубика (синий)", "Манго (желтый)",
    "Кокос + ягоды (голубой)"
], content_types=['text'])
def handle_lit_energy_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Оригинальный", "Манго", "Яблоко Киви"
], content_types=['text'])
def handle_cyberwater_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Классик черный", "Манго Лайм (зеленый)",
    "Апельсин Маракуя", "Голубика Гранат (фиолетовый)"
], content_types=['text'])
def handle_volt_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Эктив", "Бабл Гам", "Кокос", "Малина"
], content_types=['text'])
def handle_tornado_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Манго (синий)", "Экстра", "Ананас", "Апельсин"
], content_types=['text'])
def handle_gorilla_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Зеленая классика", "Обожаю Персик",
    "Черный с лимоном", "Земляничка и клюковка"
], content_types=['text'])
def handle_lipton_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Кола", "Кола Зеро", "Пепси", "7 UP", "FANTA"
], content_types=['text'])
def handle_lemonade_flavor(message):
    process_product_selection(message)

@bot.message_handler(func=lambda message: message.text in [
    "Сникерс", "Марс", "Твикс", "Баунти", "Маршмелоу",
    "Кофе Гуарана", "Орехи Манго", "Печенье"
], content_types=['text'])
def handle_snack_selection(message):
    process_product_selection(message)
def find_user_row(username):
    for row in range(2, sheet.max_row + 1):  # Начинаем с 2 строки, так как 1-я строка - заголовок
        if sheet.cell(row=row, column=1).value == username:
            return row
    return None

def process_product_selection(message):
    product = message.text
    username = message.from_user.username

    user_row = find_user_row(username)

    if user_row is not None:
        product_col = products.index(product) + 2  # products вместо unique_products

        sheet.cell(row=user_row, column=product_col, value=sheet.cell(row=user_row, column=product_col).value + 1)

        workbook.save(excel_file_path)

        bot.reply_to(message, f"Продукт {product} успешно добавлен для пользователя {username}")
    else:
        bot.reply_to(message, f"Пользователь не найден в таблице. Пожалуйста, выполните /start для начала.")

    send_welcome_with_menu(message)
def send_welcome_with_menu(message):
    bot.reply_to(message, "Привет! Я чат-бот 😇 Что ты съел сегодня?", reply_markup=main_menu_markup)

if __name__ == '__main__':
    bot.polling(none_stop=True)

